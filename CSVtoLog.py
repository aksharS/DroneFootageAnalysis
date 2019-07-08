import csv
import openpyxl

expPath = r"C:\Users\snehi\DroneFootageAnalysis\TimesTestExperimentData.xlsx"

expData = openpyxl.load_workbook(expPath)
expSheet = expData["Sheet1"]

visArray = [[]]

with open('times.csv') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    row_count = 0
    for row in csv_reader:
        visArray.insert(row_count, row)
        row_count += 1

##print(visArray)
##print(len(visArray))

for i in range(expSheet.max_row):
    if int(expSheet.cell(i + 1, 6).value) == 6:
        
        countVisRow = 0
        
        for n in range(len(visArray) - 1):
            
            print(expSheet.cell(i + 1, 2).value, int(visArray[countVisRow][1]))
            
            if expSheet.cell(i + 1, 2).value == int(visArray[countVisRow][1]):
                print(visArray[countVisRow][0])
                expSheet.cell(i + 1, 14).value = int(visArray[countVisRow][0])
            countVisRow += 1
        
        

expData.save(expPath)
