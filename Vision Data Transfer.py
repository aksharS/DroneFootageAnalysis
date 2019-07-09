import openpyxl

# Location for Vision Data Sheet
visionSheetPath = r"C:\Users\snehi\Downloads\visionSheet.xlsx"

# Location for Experiment Data Sheet
expSheetPath = r"C:\Users\snehi\Downloads\expSheet.xlsx"

# Load each workbook
visionInfo = openpyxl.load_workbook(visionSheetPath)
visionSheet = visionInfo["Sheet1"]

expInfo = openpyxl.load_workbook(expSheetPath)
expSheet = expInfo["Sheet1"]

visionData = []

# Import the data from the Vision Data Sheet into an array
def importVisionData():

    for i in range(visionSheet.max_row):
        visionData.append(int(visionSheet.cell(i + 1, 1).value))

    print("Finished Vision Data Import")

# Export the array into the experiment sheet
def exportVisionData():
    
    countVisData = 0
    for i in range(expSheet.max_row):
        if int(expSheet.cell(i + 1, 1).value) == 6:
            expSheet.cell(i + 1, 2).value = visionData[countVisData]
            countVisData += 1

    expInfo.save(expSheetPath)
    print("Finished Vision Data Export")

importVisionData()
exportVisionData()
